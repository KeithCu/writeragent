#!/usr/bin/env python3
# WriterAgent - AI Writing Assistant for LibreOffice
# Copyright (c) 2026 KeithCu (modifications and relicensing)
#
# SPDX-License-Identifier: GPL-3.0-or-later
"""Generate manual =PYTHON() serialization test XLSX for LibreOffice Calc.

Each test is a small block: input grid(s), Calc oracle, ``=PYTHON(...)``, and ``=IF(...)`` PASS/FAIL.

Input area: two side-by-side groups of up to 5 columns × 5 rows (``col_1``…``col_5`` range 1,
``col_6``…``col_10`` range 2). Single-range cases use group 1 only.

Usage (from repo root):
    python scripts/generate_serialization_spreadsheet.py
"""
from __future__ import annotations

import argparse
import io
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from tests.calc.serialization_cases import (  # noqa: E402
    SHEET_ORDER,
    SerializationCase,
    all_serialization_cases,
    case_input_grids,
)

DEFAULT_OUT = REPO_ROOT / "tests" / "fixtures"
_GROUP_COLS = 5
_GROUP_ROWS = 5
_NUM_GROUPS = 2
_MAX_INPUT_COLS = _GROUP_COLS * _NUM_GROUPS
_SHEET_NAME = "serialization_tests"
# XLSX follows Excel OOXML: comma argument separators. LibreOffice converts to locale
# (semicolon in many EU locales) on import. Semicolons in the file → Err:508 on en-US Calc.
_ARG_SEP = ","
# Display name registered in CalcAddIns.xcu; must stay uppercase in OOXML <f> cells.
_CALC_PYTHON_FN = "PYTHON"
_OOXML_PYTHON_FORMULA_RE = re.compile(r"(<f[^>]*>)(=?)python\(", re.IGNORECASE)

COL_TEST_ID = 0
COL_DESCRIPTION = 1
COL_TAGS = 2
COL_INPUT_START = 3

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_SECTION_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


@dataclass(frozen=True)
class SheetLayout:
    """Column indices for one generated sheet (fixed 2×5 input groups)."""

    max_input_cols: int = _MAX_INPUT_COLS

    @property
    def col_calc_oracle(self) -> int:
        return COL_INPUT_START + self.max_input_cols

    @property
    def col_compare(self) -> int:
        return self.col_calc_oracle + 1

    @property
    def col_expected(self) -> int:
        return self.col_compare + 1

    @property
    def col_notes(self) -> int:
        return self.col_expected + 1

    @property
    def col_python_formula(self) -> int:
        return self.col_notes + 1

    @property
    def width(self) -> int:
        return self.col_python_formula + 1

    def header(self) -> list[str]:
        col_names = [f"col_{i + 1}" for i in range(self.max_input_cols)]
        return [
            "test_id",
            "description",
            "tags",
            *col_names,
            "calc_oracle",
            "compare_pass_fail",
            "expected",
            "notes",
            "python_formula",
        ]


def group_col_start(group_index: int) -> int:
    """0-based column index where input group *group_index* begins."""
    return COL_INPUT_START + group_index * _GROUP_COLS


def calc_col_letter(col_index: int) -> str:
    """0-based column index → Calc letter (A, B, …)."""
    if col_index < 26:
        return chr(ord("A") + col_index)
    first = col_index // 26 - 1
    second = col_index % 26
    return chr(ord("A") + first) + chr(ord("A") + second)


def grid_dimensions(grid: list[list[Any]]) -> tuple[int, int]:
    if not grid:
        return 0, 0
    return len(grid), max(len(r) for r in grid)


def data_range_a1(top_row: int, nrows: int, ncols: int, *, col_start: int = COL_INPUT_START) -> str:
    """Build an A1 range for a grid anchored at *col_start* (0-based)."""
    if nrows <= 0 or ncols <= 0:
        return ""
    col_start_letter = calc_col_letter(col_start)
    col_end_letter = calc_col_letter(col_start + ncols - 1)
    if nrows == 1 and ncols == 1:
        return f"{col_start_letter}{top_row}"
    if nrows == 1:
        return f"{col_start_letter}{top_row}:{col_end_letter}{top_row}"
    if ncols == 1:
        return f"{col_start_letter}{top_row}:{col_start_letter}{top_row + nrows - 1}"
    return f"{col_start_letter}{top_row}:{col_end_letter}{top_row + nrows - 1}"


def data_ranges_for_case(case: SerializationCase, data_top: int) -> list[str]:
    """A1 refs for each input grid in *case* (group 0 left, group 1 right)."""
    ranges: list[str] = []
    for gi, grid in enumerate(case_input_grids(case)):
        nrows, ncols = grid_dimensions(grid)
        if nrows <= 0 or ncols <= 0:
            continue
        ranges.append(data_range_a1(data_top, nrows, ncols, col_start=group_col_start(gi)))
    return ranges


def _calc_oracle_formula(
    case: SerializationCase,
    data_ranges: list[str],
    block_top: int,
    *,
    primary_nrows: int,
    primary_ncols: int,
) -> str:
    if not data_ranges:
        if case.expected is not None and case.mode != "error":
            return str(case.expected)
        return ""

    primary_range = data_ranges[0]
    if case.calc_oracle == "SUM":
        if len(data_ranges) > 1:
            return f"=SUM({data_ranges[0]}{_ARG_SEP}{data_ranges[1]})"
        return f"=SUM({primary_range})"
    if case.calc_oracle == "MAX":
        return f"=MAX({primary_range})"
    if case.calc_oracle == "INDEX_FIRST":
        if primary_nrows <= 1:
            return f"=INDEX({primary_range}{_ARG_SEP}1)"
        return f"=INDEX({primary_range}{_ARG_SEP}1{_ARG_SEP}1)"
    if case.calc_oracle == "MULT2":
        return (
            f"=INDEX({primary_range}{_ARG_SEP}ROW()-{block_top - 1}{_ARG_SEP}COLUMN()-{COL_INPUT_START})*2"
        )
    if case.calc_oracle == "IDENTITY":
        return f"=INDEX({primary_range}{_ARG_SEP}ROW()-{block_top - 1}{_ARG_SEP}COLUMN()-{COL_INPUT_START})"
    if case.expected is not None and case.mode != "error":
        return str(case.expected)
    return ""


def _python_formula(case: SerializationCase, data_ranges: list[str]) -> str:
    # Single-line code only — np/pd/sp/math are auto-imported by venv_sandbox (AUTO_IMPORTS).
    code_one_line = " ".join(case.code.split())
    escaped = code_one_line.replace('"', '""')
    if case.mode == "error" or not data_ranges:
        return f'={_CALC_PYTHON_FN}("{escaped}")'
    range_part = _ARG_SEP.join(data_ranges)
    return f'={_CALC_PYTHON_FN}("{escaped}"{_ARG_SEP}{range_part})'


def cell_sheet_value(val: Any) -> int | float | str | None:
    """Calc input cell value for XLSX (native number types — not ``str(1.0)`` text)."""
    if val is None:
        return None
    if val is True:
        return 1
    if val is False:
        return 0
    if isinstance(val, (int, float, str)):
        return val
    return str(val)


def _compare_formula(layout: SheetLayout, case: SerializationCase, formula_row: int) -> str:
    calc_col = calc_col_letter(layout.col_calc_oracle)
    py_col = calc_col_letter(layout.col_python_formula)
    if case.mode == "matrix_index":
        return (
            f"matrix: Ctrl+Shift+Enter over 4x4 block in {py_col}; "
            f"each cell should match {calc_col}"
        )
    calc_cell = f"{calc_col}{formula_row}"
    py_cell = f"{py_col}{formula_row}"
    if case.mode == "error":
        return f'=IF(LEFT({py_cell}{_ARG_SEP}6)="Error:"{_ARG_SEP}"PASS"{_ARG_SEP}"FAIL")'
    if isinstance(case.expected, str):
        exp = case.expected.replace('"', '""')
        return f'=IF({py_cell}="{exp}"{_ARG_SEP}"PASS"{_ARG_SEP}"FAIL")'
    return f'=IF(ABS({calc_cell}-{py_cell})<0.001{_ARG_SEP}"PASS"{_ARG_SEP}"FAIL")'


def block_rows(layout: SheetLayout, case: SerializationCase, block_top: int) -> list[list[Any]]:
    """One test block aligned with ``layout.header()``."""
    grids = case_input_grids(case)
    max_nrows = max((grid_dimensions(g)[0] for g in grids), default=0)
    primary_nrows, primary_ncols = grid_dimensions(grids[0]) if grids else (0, 0)
    data_top = block_top + 1 if max_nrows else block_top
    data_ranges = data_ranges_for_case(case, data_top)
    width = layout.width

    header: list[Any] = [None] * width
    header[COL_TEST_ID] = case.id
    header[COL_DESCRIPTION] = case.description
    header[COL_TAGS] = ",".join(case.tags)
    header[layout.col_calc_oracle] = _calc_oracle_formula(
        case,
        data_ranges,
        data_top,
        primary_nrows=primary_nrows,
        primary_ncols=primary_ncols,
    )
    header[layout.col_compare] = _compare_formula(layout, case, block_top)
    header[layout.col_expected] = str(case.expected if case.expected is not None else case.expected_error_substr or "")
    header[layout.col_notes] = case.notes
    header[layout.col_python_formula] = _python_formula(case, data_ranges)

    lines: list[list[Any]] = [header]
    for dr in range(max_nrows):
        row: list[Any] = [None] * width
        row[COL_TEST_ID] = f"row_{dr + 1}"
        for gi, grid in enumerate(grids):
            if dr >= len(grid):
                continue
            _, ncols = grid_dimensions(grid)
            col_base = group_col_start(gi)
            for c in range(ncols):
                val = grid[dr][c]
                row[col_base + c] = cell_sheet_value(val)
        lines.append(row)

    lines.append([None] * width)
    return lines


def ordered_cases() -> list[SerializationCase]:
    """All cases in sheet order (normal → multi → mixed → grid → nan → errors)."""
    by_sheet = {s: [] for s in SHEET_ORDER}
    for case in all_serialization_cases():
        by_sheet[case.sheet].append(case)
    out: list[SerializationCase] = []
    for sheet in SHEET_ORDER:
        out.extend(by_sheet[sheet])
    return out


def _column_widths(layout: SheetLayout) -> dict[int, float]:
    widths: dict[int, float] = {
        COL_TEST_ID: 14,
        COL_DESCRIPTION: 36,
        COL_TAGS: 18,
        layout.col_calc_oracle: 22,
        layout.col_compare: 28,
        layout.col_expected: 12,
        layout.col_notes: 30,
        layout.col_python_formula: 48,
    }
    for i in range(layout.max_input_cols):
        widths[COL_INPUT_START + i] = 4
    return widths


def _wrap_columns(layout: SheetLayout) -> set[int]:
    return {COL_DESCRIPTION, layout.col_notes, layout.col_python_formula}


def build_sheet_rows(cases: list[SerializationCase]) -> tuple[SheetLayout, list[list[Any]]]:
    """Build flat rows with section band markers (``__section__:<name>`` in col 0)."""
    layout = SheetLayout()
    out: list[list[Any]] = [layout.header()]
    excel_row = 2
    prev_sheet: str | None = None
    for case in cases:
        if case.sheet != prev_sheet:
            if prev_sheet is not None:
                out.append([None] * layout.width)
                excel_row += 1
            out.append([f"__section__:{case.sheet}"] + [None] * (layout.width - 1))
            excel_row += 1
            prev_sheet = case.sheet
        block = block_rows(layout, case, excel_row)
        out.extend(block)
        excel_row += len(block)
    return layout, out


def _set_cell_value(cell, val: Any) -> None:
    """Write a value; strings starting with ``=`` become spreadsheet formulas."""
    cell.value = val


def _is_empty_cell(val: Any) -> bool:
    return val is None or val == ""


def _write_row(ws, row_idx: int, row: list[Any], layout: SheetLayout) -> None:
    if isinstance(row[0], str) and row[0].startswith("__section__:"):
        label = row[0].removeprefix("__section__:")
        # No leading "=" — Calc treats "=== … ===" as a formula (Err:510).
        cell = ws.cell(row=row_idx, column=1, value=f"[{label}]")
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=layout.width)
        return
    for col_idx, val in enumerate(row):
        if _is_empty_cell(val):
            continue
        cell = ws.cell(row=row_idx, column=col_idx + 1)
        _set_cell_value(cell, val)
        if col_idx in _wrap_columns(layout):
            cell.alignment = _WRAP


def _ensure_xlsx_python_fn_uppercase(path: Path) -> None:
    """Force ``PYTHON(`` in OOXML formula elements (openpyxl/Excel may store ``python(``)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.startswith("xl/") and info.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    text = _OOXML_PYTHON_FORMULA_RE.sub(
                        rf"\1\2{_CALC_PYTHON_FN}(",
                        text,
                    )
                    data = text.encode("utf-8")
                zout.writestr(info, data)
    path.write_bytes(buf.getvalue())


def write_combined_xlsx(path: Path, cases: list[SerializationCase]) -> SheetLayout:
    layout, rows = build_sheet_rows(cases)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME

    for row_idx, row in enumerate(rows, start=1):
        _write_row(ws, row_idx, row, layout)
        if row_idx == 1:
            for col_idx in range(layout.width):
                cell = ws.cell(row=1, column=col_idx + 1)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                if col_idx in _wrap_columns(layout):
                    cell.alignment = _WRAP

    for col_idx, width in _column_widths(layout).items():
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    _ensure_xlsx_python_fn_uppercase(path)
    return layout


def _write_readme(path: Path, layout: SheetLayout) -> None:
    group1_start = calc_col_letter(group_col_start(0))
    group1_end = calc_col_letter(group_col_start(0) + _GROUP_COLS - 1)
    group2_start = calc_col_letter(group_col_start(1))
    group2_end = calc_col_letter(group_col_start(1) + _GROUP_COLS - 1)
    oracle_col = calc_col_letter(layout.col_calc_oracle)
    cmp_col = calc_col_letter(layout.col_compare)
    py_col = calc_col_letter(layout.col_python_formula)
    path.write_text(
        f"""# =PYTHON() serialization test sheet (manual)

Open **`tests/fixtures/serialization_tests.xlsx`** in LibreOffice Calc (**File → Open**).
Calc may show an import dialog on first open — accept defaults.

## Setup

1. **Settings → Python** — set `scripting.python_venv_path` to a venv with NumPy.
2. Deploy / restart WriterAgent so `=PYTHON()` is registered.

## Layout (each test block)

| Column | Header name | Content |
|--------|-------------|---------|
| A | `test_id` | Case id (data rows use `row_N`) |
| B | `description` | What this test checks |
| C | `tags` | e.g. `split_grid`, `multi_range`, `bool` |
| {group1_start}–{group1_end} | `col_1` … `col_5` | **Range 1** (single-range cases use this group only) |
| {group2_start}–{group2_end} | `col_6` … `col_10` | **Range 2** (multi-range varargs only) |
| {oracle_col} | `calc_oracle` | Native Calc reference (`=SUM`, `=MAX`, `=INDEX`, …) |
| {cmp_col} | `compare_pass_fail` | `=IF(…;"PASS";"FAIL")` |
| … | `expected` | Expected value (reference) |
| … | `notes` | Extra hints |
| {py_col} | `python_formula` | `=PYTHON("…", range)` or `=PYTHON("…", r1, r2)` |

Each group holds up to **5 columns × 5 rows**. Multi-range cases place range 0 in group 1 and range 1 in group 2; Python receives ``data[0]``, ``data[1]``, …

Green band rows label sections: **normal** → **multi** → **mixed** → **grid** → **nan** → **errors**.

Formulas use **comma** argument separators (Excel/XLSX). LibreOffice should convert them to semicolons if your locale requires it on import. If you still see **Err:508**, check **Tools → Options → LibreOffice Calc → Formula → Separators** and edit one formula in the bar (comma ↔ semicolon) to match.

## Quick start

1. Open `tests/fixtures/serialization_tests.xlsx`.
2. Press **Ctrl+Shift+F9** (recalculate).
3. **compare_pass_fail** should show `PASS`.

Input grid (`col_1` …):
- Numeric cells are written as **real numbers** in the XLSX (not text `"1.0"`). If Calc shows them as text after import, `np.sum(data)` will fail with a string dtype error — re-run the generator or format cells as numbers.
- Bools use **`1`** / **`0`** as numbers (not `=TRUE()` formulas).
- Text `"True"` in a cell would stay a string through the wire (pickle-faithful) and break `np.sum`.

## Oracles

- **SUM** — primary; touches every numeric/logical cell. Multi-range: ``=SUM(r1,r2)``.
- **MAX** — one spot-check on 4×4 (answer 16).
- **INDEX** — first cell (text/unicode) or per-cell grid egress.

## Grid returns (grid section)

1. Select a 4×4 output area aligned with the input block.
2. Paste the formula from **python_formula** (column {py_col}) as a **matrix formula** (`Ctrl+Shift+Enter`).
3. Each cell should match **calc_oracle** (column {oracle_col}).

## Error cases (errors section)

**compare_pass_fail** should show `PASS` when **python_formula** displays `Error: …`.

## Inline code and ``float()``

Do **not** put ``float(...)`` inside the formula string (e.g. avoid ``=PYTHON("float(np.sum(data))",…)``).
Calc's formula lexer can treat ``float`` as a spreadsheet function and show **#NAME?** before Python runs.
Use ``np.sum(data)`` / ``np.max(data)`` (return values are coerced on the bridge). For longer code, put the script in a cell and use ``=PYTHON(A1, D6:G6)``.

## Regenerate

```bash
python scripts/generate_serialization_spreadsheet.py
```

Cases live in `tests/calc/serialization_cases.py`.
""",
        encoding="utf-8",
    )


def generate_all(output_dir: Path) -> None:
    cases = ordered_cases()
    layout = write_combined_xlsx(output_dir / "serialization_tests.xlsx", cases)
    _write_readme(output_dir / "serialization_tests.README.md", layout)
    print(
        f"Wrote serialization_tests.xlsx ({len(cases)} cases, "
        f"{layout.max_input_cols} input cols, {_GROUP_COLS}×{_GROUP_ROWS} per group) + README to {output_dir}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate =PYTHON() serialization test XLSX")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    generate_all(args.output_dir)


if __name__ == "__main__":
    main()
