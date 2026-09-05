# WriterAgent - AI Writing Assistant for LibreOffice
# Copyright (c) 2026 KeithCu (modifications and relicensing)
#
# SPDX-License-Identifier: GPL-3.0-or-later
"""Tests for plugin.calc.duckdb_tools (QueryFolderSqlTool)."""

from __future__ import annotations

import os
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from plugin.calc.address_utils import parse_range_string, split_sheet_prefix
from plugin.calc.duckdb_tools import (
    QueryFolderSqlTool,
    _read_sibling_office_file_as_grid,
    parse_table_source_spec,
    resolve_table_source_a1,
)
from plugin.calc.ods_cache import ODS_CACHE_DIRNAME, cache_entry_paths, write_sidecar_meta
from plugin.framework.errors import ToolExecutionError


def _mk_ctx():
    return SimpleNamespace(ctx=object(), doc=object(), doc_type="calc", active_domain="analysis")


def _write_office_fixtures(tmp_path: Path) -> tuple[Path, Path]:
    """Small xlsx/ods on disk so the host preload path sees real sibling files.

    Content is for the fixture; unit tests mock UNO used-area / read_range.
    """
    xlsx = tmp_path / "budget.xlsx"
    ods = tmp_path / "ledger.ods"
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Actuals"
        ws["C5"] = "Region"
        ws["D5"] = "Sales"
        ws["C6"] = "North"
        ws["D6"] = 100
        wb.save(xlsx)
    except ImportError:
        xlsx.write_bytes(b"PK\x03\x04xlsx")
    ods.write_bytes(b"PK\x03\x04ods")
    return xlsx, ods


def _fake_addr(start_col: int = 2, start_row: int = 4, end_col: int = 3, end_row: int = 5):
    """C5:D6 by default — data that is *not* at A1, so A1:AK2000 would pad."""
    return SimpleNamespace(StartColumn=start_col, StartRow=start_row, EndColumn=end_col, EndRow=end_row)


def _fake_sheet(name: str = "Actuals", addr=None):
    addr = addr or _fake_addr()
    cursor = SimpleNamespace(
        gotoStartOfUsedArea=lambda expand: None,
        gotoEndOfUsedArea=lambda expand: None,
        getRangeAddress=lambda: addr,
    )
    return SimpleNamespace(getName=lambda: name, createCursor=lambda: cursor)


def _fake_sheets(names: tuple[str, ...] = ("Actuals", "Summary"), addr=None):
    sheets = {n: _fake_sheet(n, addr=addr) for n in names}

    class Sheets:
        def hasByName(self, name: str) -> bool:
            return name in sheets

        def getByName(self, name: str):
            if name not in sheets:
                raise Exception(f"no sheet {name}")
            return sheets[name]

        def getByIndex(self, idx: int):
            return sheets[names[idx]]

        def getElementNames(self) -> tuple[str, ...]:
            return names

        def getCount(self) -> int:
            return len(names)

    return Sheets()


def _named_container(items: dict | None = None):
    items = items or {}

    class Container:
        def hasByName(self, name: str) -> bool:
            return name in items

        def getByName(self, name: str):
            return items[name]

        def getElementNames(self) -> tuple[str, ...]:
            return tuple(items.keys())

    return Container()


def _fake_named_range(addr, *, sheet_idx: int = 0):
    cells = SimpleNamespace(
        getRangeAddress=lambda: SimpleNamespace(
            Sheet=sheet_idx,
            StartColumn=addr.StartColumn,
            StartRow=addr.StartRow,
            EndColumn=addr.EndColumn,
            EndRow=addr.EndRow,
        )
    )
    return SimpleNamespace(getReferredCells=lambda: cells, getDataArea=lambda: None)


def _fake_database_range(addr, *, sheet_idx: int = 0):
    data_area = SimpleNamespace(
        Sheet=sheet_idx,
        StartColumn=addr.StartColumn,
        StartRow=addr.StartRow,
        EndColumn=addr.EndColumn,
        EndRow=addr.EndRow,
    )
    return SimpleNamespace(getReferredCells=lambda: None, getDataArea=lambda: data_area)


def _fake_model(sheets=None, named_ranges=None, database_ranges=None):
    sheets = sheets or _fake_sheets()
    return SimpleNamespace(
        getSheets=lambda: sheets,
        getCurrentController=lambda: None,
        NamedRanges=named_ranges if named_ranges is not None else _named_container(),
        DatabaseRanges=database_ranges if database_ranges is not None else _named_container(),
    )


def _grid_for_requested_range(range_name: str) -> list[list[dict]]:
    """Size the mock grid to the *requested* A1 range so a giant fallback is visible."""
    _prefix, bare = split_sheet_prefix(range_name)
    (start_col, start_row), (end_col, end_row) = parse_range_string(bare)
    data = {
        (2, 4): "Region",
        (3, 4): "Sales",
        (2, 5): "North",
        (3, 5): 100,
    }
    grid = []
    for row in range(start_row, end_row + 1):
        grid.append([{"value": data.get((col, row))} for col in range(start_col, end_col + 1)])
    return grid


def test_query_folder_sql_tool_basic_schema():
    t = QueryFolderSqlTool()
    assert t.name == "query_folder_sql"
    p = t.parameters
    assert "sql" in p["properties"]
    assert "sql" in p.get("required", [])
    tbl_props = p["properties"]["tables"]["additionalProperties"]["properties"]
    assert "sheet" in tbl_props
    assert "named_range" in tbl_props
    assert "range" in tbl_props
    assert "file" in tbl_props


def test_query_folder_sql_requires_sql():
    t = QueryFolderSqlTool()
    ctx = _mk_ctx()
    res = t.execute(ctx, sql="")
    assert res["status"] == "error"


@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_query_folder_sql_calls_host_with_resolved_dir(mock_resolve, mock_run, mock_exec):
    mock_resolve.return_value = "/tmp/project"
    mock_run.return_value = {"status": "ok", "helper": "query_folder_sql", "total_rows": 3}
    mock_exec.side_effect = lambda fn: fn()

    t = QueryFolderSqlTool()
    ctx = _mk_ctx()
    res = t.execute(ctx, sql="SELECT 1", files=["a.csv"])

    assert res["status"] == "ok"
    mock_resolve.assert_called()
    mock_run.assert_called_once()
    args = mock_run.call_args[0]
    assert args[1] == "/tmp/project"
    assert "SELECT 1" in args[2]


@patch("plugin.calc.duckdb_tools.os.path.isfile", return_value=True)
@patch("plugin.calc.duckdb_tools._read_sibling_office_file_as_grid")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_query_folder_sql_handles_office_files_and_sheet_hint(
    mock_resolve, mock_run, mock_exec, mock_read_office, _mock_isfile
):
    """Office files are preloaded on host; #SheetName is forwarded; CSV stays flat."""
    mock_resolve.return_value = "/tmp/project"
    mock_read_office.return_value = ("budget", [["Region", "Sales"], ["North", 100]])
    mock_run.return_value = {"status": "ok", "total_rows": 1}
    mock_exec.side_effect = lambda fn: fn()

    t = QueryFolderSqlTool()
    ctx = _mk_ctx()
    res = t.execute(ctx, sql="SELECT * FROM budget", files=["sales.csv", "budget.xlsx#Actuals"])

    assert res["status"] == "ok"
    mock_read_office.assert_called_once()
    _args, kwargs = mock_read_office.call_args
    assert _args[1].endswith("budget.xlsx")
    assert kwargs.get("sheet_hint") == "Actuals"

    call_args = mock_run.call_args
    pre = call_args.kwargs.get("preloaded")
    assert pre is not None
    assert "budget.xlsx" in pre
    assert pre["budget.xlsx"]["grid"] == [["Region", "Sales"], ["North", 100]]

    flat = call_args.kwargs.get("flat_files") or {}
    direct = call_args.kwargs.get("files")
    if direct is None and len(call_args[0]) > 3:
        direct = call_args[0][3]
    direct = direct or []
    assert any("sales.csv" in str(v) for v in (flat.values() if isinstance(flat, dict) else [])) or (
        "sales.csv" in str(direct)
    )
    assert "budget.xlsx" not in str(direct)
    assert not any("budget.xlsx" in str(v) for v in (flat.values() if isinstance(flat, dict) else []))


@patch("plugin.calc.duckdb_tools.os.path.isfile", return_value=True)
@patch("plugin.calc.duckdb_tools._read_sibling_office_file_as_grid")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_query_folder_sql_office_read_error_does_not_skip(
    mock_resolve, mock_run, mock_exec, mock_read_office, _mock_isfile
):
    """Open/sheet/empty failures must surface; SQL must not run without the table."""
    mock_resolve.return_value = "/tmp/project"
    mock_read_office.side_effect = ToolExecutionError(
        "Sibling spreadsheet 'budget.xlsx' sheet 'Nope': No sheet named 'Nope'. Available: Actuals",
        code="DUCKDB_SQL_ERROR",
    )
    mock_exec.side_effect = lambda fn: fn()

    t = QueryFolderSqlTool()
    res = t.execute(_mk_ctx(), sql="SELECT * FROM budget", files=["budget.xlsx#Nope"])

    assert res["status"] == "error"
    assert "budget.xlsx" in res["message"]
    assert "Nope" in res["message"]
    mock_run.assert_not_called()


@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_read_sibling_used_range_skips_giant_padding(mock_open, mock_read, _mock_close, tmp_path):
    """Used-area C5:D6 must not ship A1:AK2000 / A1:AZ5000 empty padding."""
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_read.side_effect = _grid_for_requested_range

    _tbl, grid = _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")

    requested = mock_read.call_args[0][0]
    assert "AK2000" not in requested
    assert "AZ5000" not in requested
    _prefix, bare = split_sheet_prefix(requested)
    assert _prefix == "Actuals"
    assert bare == "C5:D6"
    assert len(grid) == 2
    assert all(len(row) == 2 for row in grid)
    assert grid[0] == ["Region", "Sales"]
    assert grid[1] == ["North", 100]


@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_read_sibling_ods_used_range(mock_open, mock_read, _mock_close, tmp_path):
    _xlsx, ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_read.side_effect = _grid_for_requested_range

    _tbl, grid = _read_sibling_office_file_as_grid(object(), str(ods))

    requested = mock_read.call_args[0][0]
    assert "AK2000" not in requested
    assert len(grid) == 2
    assert len(grid[0]) == 2


@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_read_sibling_missing_sheet_hint_errors(mock_open, mock_read, _mock_close, tmp_path):
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)

    try:
        _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Nope")
    except ToolExecutionError as exc:
        assert "budget.xlsx" in str(exc)
        assert "Nope" in str(exc)
        assert "Actuals" in str(exc)
    else:
        raise AssertionError("expected ToolExecutionError for missing sheet hint")
    mock_read.assert_not_called()


@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_read_sibling_open_failure_errors(mock_open, _mock_close, tmp_path):
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (None, None, "Failed to open /tmp/budget.xlsx", False)

    try:
        _read_sibling_office_file_as_grid(object(), str(xlsx))
    except ToolExecutionError as exc:
        assert "budget.xlsx" in str(exc)
        assert "Failed to open" in str(exc)
    else:
        raise AssertionError("expected ToolExecutionError when LibreOffice open fails")


@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_read_sibling_empty_used_range_errors(mock_open, mock_read, _mock_close, tmp_path):
    xlsx, _ods = _write_office_fixtures(tmp_path)
    empty_addr = _fake_addr(0, 0, 0, 0)
    mock_open.return_value = (_fake_model(sheets=_fake_sheets(addr=empty_addr)), "calc", None, True)
    mock_read.return_value = [[{"value": None}]]

    try:
        _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")
    except ToolExecutionError as exc:
        msg = str(exc)
        assert "budget.xlsx" in msg
        assert "Actuals" in msg
        assert "empty" in msg.lower()
    else:
        raise AssertionError("expected ToolExecutionError for empty used range")


@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_preload_path_ships_used_range_grid(
    mock_resolve, mock_run, mock_exec, _mock_close, mock_open, mock_read, tmp_path
):
    """End-to-end host preload: fixture files + used-range, no giant padding."""
    xlsx, ods = _write_office_fixtures(tmp_path)
    (tmp_path / "sales.csv").write_text("region,amount\nNorth,1\n")
    mock_resolve.return_value = str(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_read.side_effect = _grid_for_requested_range
    mock_run.return_value = {"status": "ok", "total_rows": 1}
    mock_exec.side_effect = lambda fn: fn()

    t = QueryFolderSqlTool()
    res = t.execute(
        _mk_ctx(),
        sql="SELECT * FROM budget",
        files=["sales.csv", "budget.xlsx#Actuals", "ledger.ods"],
    )
    assert res["status"] == "ok"

    pre = mock_run.call_args.kwargs.get("preloaded") or {}
    assert "budget.xlsx" in pre
    assert "ledger.ods" in pre
    budget_grid = pre["budget.xlsx"]["grid"]
    assert len(budget_grid) == 2
    assert len(budget_grid[0]) == 2
    assert budget_grid[0] == ["Region", "Sales"]
    for requested in (c[0][0] for c in mock_read.call_args_list):
        assert "AK2000" not in requested
        assert "AZ5000" not in requested


@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_preload_path_bad_sheet_hint_errors(
    mock_resolve, mock_run, mock_exec, _mock_close, mock_open, mock_read, tmp_path
):
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_resolve.return_value = str(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_exec.side_effect = lambda fn: fn()

    t = QueryFolderSqlTool()
    res = t.execute(_mk_ctx(), sql="SELECT * FROM budget", files=[f"{xlsx.name}#MissingSheet"])

    assert res["status"] == "error"
    assert "budget.xlsx" in res["message"]
    assert "MissingSheet" in res["message"]
    mock_run.assert_not_called()
    mock_read.assert_not_called()


def test_parse_table_source_spec_identities():
    sheet = parse_table_source_spec({"sheet": "Sales_Analytics", "headers": False})
    assert sheet["kind"] == "sheet"
    assert sheet["sheet"] == "Sales_Analytics"
    assert sheet["range"] is None
    assert sheet["headers"] is False

    named = parse_table_source_spec({"named_range": "SalesData"})
    assert named["kind"] == "named_range"
    assert named["named_range"] == "SalesData"

    frozen = parse_table_source_spec({"range": "Sales.A1:F500"})
    assert frozen["kind"] == "range"
    assert frozen["range"] == "Sales.A1:F500"

    legacy = parse_table_source_spec("Costs.A1:D200")
    assert legacy["kind"] == "range"
    assert legacy["range"] == "Costs.A1:D200"

    sibling = parse_table_source_spec({"file": "budget.xlsx#Sales"})
    assert sibling["kind"] == "sheet"
    assert sibling["sheet"] == "Sales"
    assert sibling["file"] == "budget.xlsx"


def test_parse_table_source_spec_rejects_mixed_or_empty():
    try:
        parse_table_source_spec({"sheet": "Sales", "range": "A1:B2"})
    except ToolExecutionError as exc:
        assert "exactly one" in str(exc)
    else:
        raise AssertionError("mixed sheet+range must fail")

    try:
        parse_table_source_spec({})
    except ToolExecutionError as exc:
        assert "sheet" in str(exc)
        assert "named_range" in str(exc)
    else:
        raise AssertionError("empty spec must fail")


def test_sheet_identity_rereads_used_range_not_cached_a1():
    """Same {sheet} spec must pick up a grown used-range — catalog is not expanded A1."""
    addr = _fake_addr(2, 4, 3, 5)
    model = _fake_model(sheets=_fake_sheets(names=("Actuals",), addr=addr))
    parsed = parse_table_source_spec({"sheet": "Actuals"})

    first = resolve_table_source_a1(model, parsed)
    assert first.endswith("C5:D6")
    assert parsed["range"] is None
    assert parsed["sheet"] == "Actuals"

    addr.EndRow = 7
    second = resolve_table_source_a1(model, parsed)
    assert second.endswith("C5:D8")
    assert parsed["range"] is None


def test_named_range_identity_follows_referred_bounds():
    addr = _fake_addr(2, 4, 3, 5)
    nr = _fake_named_range(addr)
    model = _fake_model(
        sheets=_fake_sheets(names=("Actuals",), addr=addr),
        named_ranges=_named_container({"SalesData": nr}),
    )
    parsed = parse_table_source_spec({"named_range": "SalesData"})

    first = resolve_table_source_a1(model, parsed)
    assert first.endswith("C5:D6")
    addr.EndRow = 8
    second = resolve_table_source_a1(model, parsed)
    assert second.endswith("C5:D9")
    assert parsed["named_range"] == "SalesData"
    assert parsed["range"] is None


def test_database_range_identity_uses_data_area():
    addr = _fake_addr(0, 0, 1, 3)
    dbr = _fake_database_range(addr)
    model = _fake_model(
        sheets=_fake_sheets(names=("Actuals",), addr=addr),
        database_ranges=_named_container({"CostDB": dbr}),
    )
    parsed = parse_table_source_spec({"named_range": "CostDB"})
    qualified = resolve_table_source_a1(model, parsed)
    assert qualified.endswith("A1:B4")


def test_named_range_identity_missing_errors():
    model = _fake_model()
    parsed = parse_table_source_spec({"named_range": "Nope"})
    try:
        resolve_table_source_a1(model, parsed)
    except ToolExecutionError as exc:
        assert "Nope" in str(exc)
    else:
        raise AssertionError("missing named range must fail loud")


def test_frozen_range_identity_is_not_used_range():
    addr = _fake_addr(2, 4, 3, 5)
    model = _fake_model(sheets=_fake_sheets(names=("Actuals",), addr=addr))
    parsed = parse_table_source_spec({"range": "Actuals.A1:B2"})
    assert resolve_table_source_a1(model, parsed) == "Actuals.A1:B2"


@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_execute_sheet_identity_preloads_used_range(mock_resolve, mock_run, mock_exec, mock_read):
    mock_resolve.return_value = "/tmp/project"
    mock_run.return_value = {"status": "ok", "total_rows": 1}
    mock_exec.side_effect = lambda fn: fn()
    mock_read.side_effect = _grid_for_requested_range

    ctx = _mk_ctx()
    ctx.doc = _fake_model()
    res = QueryFolderSqlTool().execute(
        ctx,
        sql="SELECT * FROM sales",
        tables={"sales": {"sheet": "Actuals"}},
    )
    assert res["status"] == "ok"
    requested = mock_read.call_args[0][0]
    assert "C5:D6" in requested
    assert "AK2000" not in requested
    pre = mock_run.call_args.kwargs.get("preloaded") or {}
    assert "sales" in pre
    assert pre["sales"]["grid"][0] == ["Region", "Sales"]


@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_execute_named_range_identity(mock_resolve, mock_run, mock_exec, mock_read):
    mock_resolve.return_value = "/tmp/project"
    mock_run.return_value = {"status": "ok"}
    mock_exec.side_effect = lambda fn: fn()
    mock_read.side_effect = _grid_for_requested_range
    addr = _fake_addr()
    ctx = _mk_ctx()
    ctx.doc = _fake_model(named_ranges=_named_container({"SalesData": _fake_named_range(addr)}))

    res = QueryFolderSqlTool().execute(
        ctx,
        sql="SELECT * FROM sales",
        tables={"sales": {"named_range": "SalesData"}},
    )
    assert res["status"] == "ok"
    requested = mock_read.call_args[0][0]
    assert "C5:D6" in requested
    pre = mock_run.call_args.kwargs.get("preloaded") or {}
    assert "sales" in pre


@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_execute_frozen_range_stays_absolute(mock_resolve, mock_run, mock_exec, mock_read):
    mock_resolve.return_value = "/tmp/project"
    mock_run.return_value = {"status": "ok"}
    mock_exec.side_effect = lambda fn: fn()
    mock_read.side_effect = _grid_for_requested_range
    ctx = _mk_ctx()
    ctx.doc = _fake_model()

    res = QueryFolderSqlTool().execute(
        ctx,
        sql="SELECT * FROM sales",
        tables={"sales": {"range": "Actuals.A1:B2"}},
    )
    assert res["status"] == "ok"
    requested = mock_read.call_args[0][0]
    _prefix, bare = split_sheet_prefix(requested)
    assert _prefix == "Actuals"
    assert bare == "A1:B2"


@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_execute_missing_named_range_does_not_run_sql(mock_resolve, mock_run, mock_exec):
    mock_resolve.return_value = "/tmp/project"
    mock_exec.side_effect = lambda fn: fn()
    ctx = _mk_ctx()
    ctx.doc = _fake_model()

    res = QueryFolderSqlTool().execute(
        ctx,
        sql="SELECT * FROM sales",
        tables={"sales": {"named_range": "MissingName"}},
    )
    assert res["status"] == "error"
    assert "MissingName" in res["message"]
    mock_run.assert_not_called()


@patch("plugin.calc.duckdb_tools.os.path.isfile", return_value=True)
@patch("plugin.calc.duckdb_tools._read_sibling_office_file_as_grid")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_execute_tables_file_sheet_identity(
    mock_resolve, mock_run, mock_exec, mock_read_office, _mock_isfile
):
    """tables={name: {file, sheet}} uses the sibling used-range path; catalog name is the key."""
    mock_resolve.return_value = "/tmp/project"
    mock_read_office.return_value = ("budget", [["Region", "Sales"], ["North", 100]])
    mock_run.return_value = {"status": "ok"}
    mock_exec.side_effect = lambda fn: fn()

    res = QueryFolderSqlTool().execute(
        _mk_ctx(),
        sql="SELECT * FROM sales",
        tables={"sales": {"file": "budget.xlsx", "sheet": "Actuals"}},
    )
    assert res["status"] == "ok"
    _args, kwargs = mock_read_office.call_args
    assert _args[1].endswith("budget.xlsx")
    assert kwargs.get("sheet_hint") == "Actuals"
    pre = mock_run.call_args.kwargs.get("preloaded") or {}
    assert "sales" in pre
    assert pre["sales"]["grid"][0] == ["Region", "Sales"]


def _fake_export_writes_ods(model, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(b"PK\x03\x04cached-ods")


@patch("plugin.calc.duckdb_tools._source_is_open_workbook", return_value=False)
@patch("plugin.calc.duckdb_tools._export_model_to_cached_ods", side_effect=_fake_export_writes_ods)
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_xlsx_cache_miss_then_hit_mtime_invalidates(
    mock_open, mock_read, _mock_close, _mock_export, _mock_open_wb, tmp_path
):
    """First XLSX open writes cache; second opens the ODS; mtime change misses."""
    xlsx, _ods = _write_office_fixtures(tmp_path)
    opened: list[str] = []

    def _open(_ctx, path):
        opened.append(os.path.normpath(str(path)))
        return (_fake_model(), "calc", None, True)

    mock_open.side_effect = _open
    mock_read.side_effect = _grid_for_requested_range

    _tbl, grid1 = _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")
    assert grid1[0] == ["Region", "Sales"]
    assert opened[0] == os.path.normpath(str(xlsx))
    paths = cache_entry_paths(str(xlsx))
    assert paths is not None
    cached_ods, meta_path = paths
    assert cached_ods.is_file()
    assert meta_path.is_file()

    _tbl, grid2 = _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")
    assert grid2[0] == ["Region", "Sales"]
    assert opened[1] == os.path.normpath(str(cached_ods))

    st = os.stat(xlsx)
    os.utime(xlsx, ns=(st.st_atime_ns, st.st_mtime_ns + 1_000_000_000))
    _tbl, grid3 = _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")
    assert grid3[0] == ["Region", "Sales"]
    assert opened[2] == os.path.normpath(str(xlsx))


@patch("plugin.calc.duckdb_tools._source_is_open_workbook", return_value=False)
@patch("plugin.calc.duckdb_tools._export_model_to_cached_ods", side_effect=_fake_export_writes_ods)
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_ods_source_skips_cache(mock_open, mock_read, _mock_close, mock_export, _mock_open_wb, tmp_path):
    _xlsx, ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_read.side_effect = _grid_for_requested_range

    _read_sibling_office_file_as_grid(object(), str(ods))

    mock_open.assert_called_once()
    assert os.path.normpath(mock_open.call_args[0][1]) == os.path.normpath(str(ods))
    mock_export.assert_not_called()
    assert not (tmp_path / ODS_CACHE_DIRNAME).exists()


@patch("plugin.calc.ods_cache.ods_cache_enabled", return_value=False)
@patch("plugin.calc.duckdb_tools._source_is_open_workbook", return_value=False)
@patch("plugin.calc.duckdb_tools._export_model_to_cached_ods")
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_xlsx_skips_cache_when_disabled(
    mock_open, mock_read, _mock_close, mock_export, _mock_open_wb, _enabled, tmp_path
):
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, True)
    mock_read.side_effect = _grid_for_requested_range

    _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")

    assert os.path.normpath(mock_open.call_args[0][1]) == os.path.normpath(str(xlsx))
    mock_export.assert_not_called()
    assert not (tmp_path / ODS_CACHE_DIRNAME).exists()


@patch("plugin.calc.duckdb_tools._source_is_open_workbook", return_value=True)
@patch("plugin.calc.duckdb_tools._export_model_to_cached_ods")
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
def test_live_workbook_skips_cache(mock_open, mock_read, _mock_close, mock_export, _mock_open_wb, tmp_path):
    """Already-open workbook: skip cache hit and do not storeToURL (opened_flag=False)."""
    xlsx, _ods = _write_office_fixtures(tmp_path)
    mock_open.return_value = (_fake_model(), "calc", None, False)
    mock_read.side_effect = _grid_for_requested_range

    _read_sibling_office_file_as_grid(object(), str(xlsx), sheet_hint="Actuals")

    assert os.path.normpath(mock_open.call_args[0][1]) == os.path.normpath(str(xlsx))
    mock_export.assert_not_called()
    _mock_close.assert_not_called()


@patch("plugin.calc.duckdb_tools._source_is_open_workbook", return_value=False)
@patch("plugin.calc.duckdb_tools.close_document_research_document")
@patch("plugin.calc.inspector.CellInspector.read_range")
@patch("plugin.calc.duckdb_tools.open_document_for_read")
@patch("plugin.calc.duckdb_tools.execute_on_main_thread")
@patch("plugin.scripting.client.run_folder_sql")
@patch("plugin.calc.duckdb_tools.resolve_listing_directory")
def test_sql_path_uses_cache_on_second_query(
    mock_resolve, mock_run, mock_exec, mock_open, mock_read, _mock_close, _mock_open_wb, tmp_path
):
    """query_folder_sql still preloads the used-range grid when the open is a cache hit."""
    xlsx, _ods = _write_office_fixtures(tmp_path)
    (tmp_path / "sales.csv").write_text("region,amount\nNorth,1\n")
    mock_resolve.return_value = str(tmp_path)
    mock_run.return_value = {"status": "ok", "helper": "query_folder_sql", "total_rows": 1}
    mock_exec.side_effect = lambda fn: fn()
    mock_read.side_effect = _grid_for_requested_range

    opened: list[str] = []

    def _open(_ctx, path):
        opened.append(os.path.normpath(str(path)))
        return (_fake_model(), "calc", None, True)

    mock_open.side_effect = _open

    # Seed a valid cache entry so the SQL open is a hit.
    paths = cache_entry_paths(str(xlsx))
    assert paths is not None
    cached_ods, meta_path = paths
    cached_ods.parent.mkdir(parents=True, exist_ok=True)
    cached_ods.write_bytes(b"PK\x03\x04cached-ods")
    write_sidecar_meta(meta_path, str(xlsx))

    t = QueryFolderSqlTool()
    res = t.execute(
        _mk_ctx(),
        sql="SELECT * FROM budget",
        files=["sales.csv", "budget.xlsx#Actuals"],
    )
    assert res["status"] == "ok"
    assert opened == [os.path.normpath(str(cached_ods))]
    pre = mock_run.call_args.kwargs.get("preloaded") or {}
    assert "budget.xlsx" in pre
    assert pre["budget.xlsx"]["grid"] == [["Region", "Sales"], ["North", 100]]
    flat = mock_run.call_args.kwargs.get("flat_files") or {}
    assert any(str(v).endswith("sales.csv") for v in flat.values())
