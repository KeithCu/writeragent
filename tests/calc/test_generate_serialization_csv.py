# WriterAgent - AI Writing Assistant for LibreOffice
# Copyright (c) 2026 KeithCu (modifications and relicensing)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

from __future__ import annotations

import importlib.util
import re
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT))

from tests.calc.serialization_cases import all_serialization_cases  # noqa: E402

_GEN_PATH = REPO_ROOT / "scripts" / "generate_serialization_test_csv.py"
_spec = importlib.util.spec_from_file_location("generate_serialization_test_csv", _GEN_PATH)
assert _spec and _spec.loader
gen = importlib.util.module_from_spec(_spec)
sys.modules[_spec.name] = gen
_spec.loader.exec_module(gen)


def _case_by_id(case_id: str):
    return next(c for c in all_serialization_cases() if c.id == case_id)


def test_layout_python_formula_is_last_column():
    layout = gen.SheetLayout(max_input_cols=10)
    assert layout.col_python_formula == layout.col_notes + 1
    assert layout.header()[-1] == "python_formula"
    assert "python_code" not in layout.header()
    assert layout.header()[3] == "col_1"


def test_max_input_cols_covers_row_10_and_grid_2x5():
    cases = gen.ordered_cases()
    max_cols = gen.max_input_cols_for_cases(cases)
    assert max_cols >= 10
    _, ncols_2x5 = gen.grid_dimensions(_case_by_id("grid_2x5_sum").input_grid)
    _, ncols_row10 = gen.grid_dimensions(_case_by_id("row_10_sum").input_grid)
    assert ncols_2x5 == 5
    assert ncols_row10 == 10
    assert max_cols >= ncols_2x5
    assert max_cols >= ncols_row10


def test_generated_blocks_write_all_input_cells():
    cases = gen.ordered_cases()
    layout = gen.SheetLayout(max_input_cols=gen.max_input_cols_for_cases(cases))
    row = 2
    for case in cases:
        block = gen.block_rows(layout, case, row)
        nrows, ncols = gen.grid_dimensions(case.input_grid)
        if nrows == 0:
            row += len(block)
            continue
        for dr in range(nrows):
            data_row = block[1 + dr]
            for c in range(ncols):
                expected = case.input_grid[dr][c]
                cell = data_row[gen.COL_INPUT_START + c]
                if expected is None:
                    assert cell in (None, ""), f"{case.id} row {dr + 1} col {c + 1} should be empty"
                else:
                    assert cell not in (None, ""), f"{case.id} row {dr + 1} col {c + 1} missing in sheet row"
        data_top = row + 1 if nrows else row
        expected_range = gen.data_range_a1(data_top, nrows, ncols)
        header = block[0]
        py_formula = header[layout.col_python_formula]
        if case.mode != "error" and nrows:
            assert expected_range in py_formula, f"{case.id}: formula missing range {expected_range!r}"
        assert py_formula.startswith("=PYTHON")
        assert not py_formula.startswith("#")
        row += len(block)


def test_bool_cells_use_numeric_one_zero():
    cases = gen.ordered_cases()
    layout = gen.SheetLayout(max_input_cols=gen.max_input_cols_for_cases(cases))
    for case_id in ("bool_true", "bool_false", "bool_col_11_sum"):
        case = _case_by_id(case_id)
        block = gen.block_rows(layout, case, 2)
        for line in block[1:]:
            row_id = line[gen.COL_TEST_ID]
            if not isinstance(row_id, str) or not row_id.startswith("row_"):
                continue
            val = line[gen.COL_INPUT_START]
            if val not in (None, ""):
                assert val in (0, 1), f"{case_id}: expected 0/1 ints, got {val!r}"


def test_xlsx_input_cells_are_numeric(tmp_path: Path) -> None:
    gen.generate_all(tmp_path)
    wb = load_workbook(tmp_path / "serialization_tests.xlsx", read_only=True, data_only=True)
    ws = wb["serialization_tests"]
    case = _case_by_id("scalar_row_sum")
    found = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value != case.id:
            continue
        data_row = row + 1
        val = ws.cell(row=data_row, column=gen.COL_INPUT_START + 1).value
        assert isinstance(val, (int, float)), f"scalar_row_sum row_1 col_1 should be numeric, got {type(val).__name__!r}: {val!r}"
        found = True
        break
    assert found
    wb.close()


def test_matrix_index_formula_has_two_args_only():
    layout = gen.SheetLayout(max_input_cols=gen.max_input_cols_for_cases(gen.ordered_cases()))
    for case_id in ("grid_return_double", "grid_return_identity"):
        case = _case_by_id(case_id)
        block = gen.block_rows(layout, case, 2)
        py_formula = block[0][layout.col_python_formula]
        nrows, ncols = gen.grid_dimensions(case.input_grid)
        data_top = 3 if nrows else 2
        expected_range = gen.data_range_a1(data_top, nrows, ncols)
        assert expected_range in py_formula, f"{case_id}: missing range {expected_range!r}"
        assert "ROW()" not in py_formula, f"{case_id}: must not use third ROW() arg (Err:504)"
        assert py_formula.startswith("=PYTHON")
        inner = py_formula.split("(", 1)[1]
        assert inner.count(")") >= 1
        args_part = inner.rsplit(")", 1)[0]
        assert args_part.count(",") == 1, f"{case_id}: expected 2-arg PYTHON, got {py_formula!r}"


def test_wide_case_formula_range_column_count():
    layout = gen.SheetLayout(max_input_cols=gen.max_input_cols_for_cases(gen.ordered_cases()))
    case = _case_by_id("row_10_sum")
    block = gen.block_rows(layout, case, 2)
    py_formula = block[0][layout.col_python_formula]
    m = re.search(r"[,;]([A-Z]+\d+(?::[A-Z]+\d+)?)\)", py_formula)
    assert m, py_formula
    a1 = m.group(1)
    assert "M" in a1 or a1.count(":") == 0


def test_generate_xlsx_smoke(tmp_path: Path) -> None:
    gen.generate_all(tmp_path)
    xlsx = tmp_path / "serialization_tests.xlsx"
    assert xlsx.is_file()
    wb = load_workbook(xlsx, read_only=True, data_only=False)
    ws = wb["serialization_tests"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 20) if ws.cell(row=1, column=c).value]
    assert headers[-1] == "python_formula"
    assert "python_code" not in headers
    assert "col_1" in headers

    case = _case_by_id("scalar_single_cell")
    layout = gen.SheetLayout(max_input_cols=gen.max_input_cols_for_cases(gen.ordered_cases()))
    py_col = layout.col_python_formula + 1
    found_id = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == case.id:
            found_id = True
            formula = ws.cell(row=row, column=py_col).value
            assert isinstance(formula, str)
            assert formula.startswith("=PYTHON")
            break
    assert found_id

    import zipfile

    with zipfile.ZipFile(xlsx) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode()
    assert re.search(r"<f[^>]*>PYTHON\(", sheet_xml), "OOXML formulas must use PYTHON("
    assert not re.search(r"<f[^>]*>python\(", sheet_xml), "OOXML must not use lowercase python("

    section_ok = False
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val == "[normal]":
            section_ok = True
            break
    assert section_ok
    wb.close()
