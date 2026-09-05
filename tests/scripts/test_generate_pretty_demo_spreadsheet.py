# WriterAgent - AI Writing Assistant for LibreOffice
# Copyright (c) 2026 KeithCu (modifications and relicensing)
#
# SPDX-License-Identifier: GPL-3.0-or-later
"""Unit tests for scripts/generate_pretty_demo_spreadsheet.py."""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

from scripts.generate_pretty_demo_spreadsheet import (
    CALC_PYTHON_ADDIN_FN,
    RESULTS_PY_CODE_MAX_LEN,
    SALES_RANGE_ODS_CROSS,
    SALES_ZIPS_BY_REGION,
    SQL_RESULTS_SPILL_GUTTER_COLS,
    SQL_RESULTS_SPILL_GUTTER_ROWS,
    SQL_SALES_BY_REGION_CATEGORY,
    SQL_SALES_ZIP_INCOME_JOIN,
    ZIP_INCOME_CSV_NAME,
    ZIP_INCOME_FIXTURE,
    _ODS_SHEET_COLUMNS,
    _scenario_result_formula,
    build_ods_showcase,
    build_xlsx_showcase,
    duckdb_sql_from_cell_code,
    get_sales_dataset,
    ods_formula,
    sql_demo_scenarios,
    sql_query_lines,
    sql_results_gutter_rows,
    write_zip_income_csv,
)

# Tokens that must live in sheet cells, never inside the RESULTS formula string.
_SQL_EMBED_MARKERS = (
    "SUM(Revenue)",
    "SUM(Ad_Spend)",
    "GROUP BY",
    "COUNT(*)",
    "NULLIF",
    "FROM sales",
    "FROM marketing",
)

FIXTURE_ODS = Path(__file__).resolve().parents[1] / "fixtures" / "python_showcase_demo.ods"
_NESTED_SHEET_REF = re.compile(r"\[\$[A-Za-z0-9_]*\[\$")


def _ods_content_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read("content.xml").decode("utf-8")


def test_ods_formula_sales_analytics_range_not_rematched() -> None:
    out = ods_formula('=PY("x"; Sales_Analytics.A5:I40)')
    assert "[$Sales_Analytics.A5:.I40]" in out
    assert _NESTED_SHEET_REF.search(out) is None
    assert "[$S[$ales_Analytics" not in out
    assert out.startswith(f"of:={CALC_PYTHON_ADDIN_FN}(")


def test_ods_formula_forecasting_range_not_rematched() -> None:
    out = ods_formula('=PY("y"; Forecasting.A1:B2)')
    assert "[$Forecasting.A1:.B2]" in out
    assert _NESTED_SHEET_REF.search(out) is None
    assert "[$F[$orecasting" not in out


def test_ods_formula_same_sheet_range_and_cell() -> None:
    out = ods_formula('=PY("sum(r[7] for r in data[1:])"; A5:I40; F46)')
    assert "[.A5:.I40]" in out
    assert "[.F46]" in out
    assert "A5:I40" not in out
    assert out.startswith(f"of:={CALC_PYTHON_ADDIN_FN}(")


def test_ods_formula_python_wrapper_and_cross_sheet_cell() -> None:
    out = ods_formula('=PYTHON("x"; Sales_Analytics.F47)')
    assert out.startswith(f"of:={CALC_PYTHON_ADDIN_FN}(")
    assert "[$Sales_Analytics.F47]" in out
    assert _NESTED_SHEET_REF.search(out) is None


def test_ods_formula_leaves_non_formula_text() -> None:
    assert ods_formula("plain") == "plain"


def test_ods_formula_statistics_ml_range_not_rematched() -> None:
    out = ods_formula('=PY("x"; Statistics_ML.A5:G25)')
    assert "[$Statistics_ML.A5:.G25]" in out
    assert _NESTED_SHEET_REF.search(out) is None


def _quoted_formula_payload(formula: str) -> str:
    """First ``"…"`` argument of ``=FN("code", range)``. Exactly one pair of quotes."""
    assert formula.startswith("="), formula
    first = formula.index('"')
    last = formula.rindex('"')
    assert first < last, formula
    # A premature " (e.g. Python """) would add extra quotes and leak SQL
    # tokens like SUM(...) into the formula — Calc Err:508.
    assert formula.count('"') == 2, formula
    return formula[first + 1 : last]


def test_sales_dataset_has_zip_column_from_real_zctas() -> None:
    grid = get_sales_dataset()
    assert grid[0][-1] == "ZIP"
    assert grid[0].index("Revenue") == 7
    zips = {str(row[-1]) for row in grid[1:]}
    allowed = {z for group in SALES_ZIPS_BY_REGION.values() for z in group}
    assert zips <= allowed
    assert zips  # at least one assigned


def test_sql_demo_scenarios_include_sheet_and_join() -> None:
    kinds = {s["kind"] for s in sql_demo_scenarios()}
    assert kinds == {"sheet_sales", "sheet_marketing", "join_zip"}
    join = next(s for s in sql_demo_scenarios() if s["kind"] == "join_zip")
    assert "zip_income" in join["sql"]
    assert join["sql"] == SQL_SALES_ZIP_INCOME_JOIN
    assert "ZIP" in join["sql"] or "zip" in join["sql"]


def test_zip_income_csv_is_fuller_acs_extract_and_covers_sales_zips() -> None:
    assert ZIP_INCOME_FIXTURE.is_file()
    lines = ZIP_INCOME_FIXTURE.read_text(encoding="utf-8").splitlines()
    assert lines[0].startswith("zip,median_household_income")
    # Fuller ZCTA extract (not a 40-row toy); ACS has ~33k ZCTAs.
    assert len(lines) > 5000
    csv_zips = {line.split(",", 1)[0] for line in lines[1:] if line.strip()}
    sales_zips = {z for group in SALES_ZIPS_BY_REGION.values() for z in group}
    missing = sales_zips - csv_zips
    assert not missing, f"sales ZIPs missing from ACS extract: {missing}"


def test_write_zip_income_csv_copies_sibling(tmp_path: Path) -> None:
    dest = write_zip_income_csv(tmp_path)
    assert dest.name == ZIP_INCOME_CSV_NAME
    assert dest.is_file()
    assert dest.stat().st_size == ZIP_INCOME_FIXTURE.stat().st_size


def test_generated_ods_has_sql_duckdb_sheet(tmp_path: Path) -> None:
    out = tmp_path / "python_showcase_demo.ods"
    build_ods_showcase(out)
    write_zip_income_csv(tmp_path)
    assert out.is_file()
    assert (tmp_path / ZIP_INCOME_CSV_NAME).is_file()

    from odf.opendocument import load
    from odf.table import Table
    from odf.text import P

    doc = load(str(out))
    names = [str(t.getAttribute("name")) for t in doc.spreadsheet.getElementsByType(Table)]
    assert "Sales_Analytics" in names
    assert "SQL_DuckDB" in names
    sql_sheet = next(t for t in doc.spreadsheet.getElementsByType(Table) if t.getAttribute("name") == "SQL_DuckDB")
    text = "\n".join(str(p) for p in sql_sheet.getElementsByType(P))
    assert "zip_income" in text
    assert "GROUP BY" in text or "Region" in text
    _assert_ods_results_unmerged_with_clearance(sql_sheet)


def _assert_ods_formulas_and_layout(xml: str) -> None:
    sales_of = f"[${SALES_RANGE_ODS_CROSS.replace(':', ':.')}]"
    assert sales_of in xml
    assert "[$Forecasting.A5:.E41]" in xml
    assert "[$S[$ales_Analytics" not in xml
    assert "[$F[$orecasting" not in xml
    assert "[$S[$tatistics_ML" not in xml
    assert _NESTED_SHEET_REF.search(xml) is None
    assert xml.count("<table:table-column") >= sum(len(cols) for cols in _ODS_SHEET_COLUMNS.values())
    assert "style:column-width" in xml
    assert "style:row-height" in xml
    assert "SQL_DuckDB" in xml
    # SQL text is cell content; RESULTS formulas are short OpenFormula runners.
    assert "SUM(Revenue)" in xml
    assert "SUM(Ad_Spend)" in xml
    formula_attrs = re.findall(r'table:formula="([^"]*)"', xml)
    # ``data[1]`` also appears in Forecasting CGR; pin RESULTS on the DuckDB runner.
    results = [f for f in formula_attrs if "con.sql(sql)" in f]
    assert len(results) == 2
    for formula in results:
        assert "data[1]" in formula
        assert "result=con.sql(sql).df()" in formula
        assert ".tolist()" not in formula
        assert "[.A" in formula  # explicit SQL cell/range, not magic-above
        assert len(formula) < 400
        for marker in _SQL_EMBED_MARKERS:
            assert marker not in formula, (marker, formula)
    # Live RESULTS formula cells must be unmerged (no number-columns-spanned).
    for m in re.finditer(r"<table:table-cell\b([^>]*)>", xml):
        attrs = m.group(1)
        if "con.sql(sql)" not in attrs:
            continue
        assert "number-columns-spanned" not in attrs, attrs


def _ods_row_is_empty(row: object) -> bool:
    from odf.table import TableCell
    from odf.text import P

    cells = row.getElementsByType(TableCell)  # type: ignore[attr-defined]
    if not cells:
        return True
    for cell in cells:
        if cell.getAttribute("formula"):
            return False
        texts = [str(p) for p in cell.getElementsByType(P) if str(p).strip()]
        if texts:
            return False
    return True


def _assert_ods_results_unmerged_with_clearance(sql_sheet: object) -> None:
    """Live ODS RESULTS cells have no column span; next 15 rows are empty."""
    from odf.table import TableCell, TableRow

    rows = list(sql_sheet.getElementsByType(TableRow))  # type: ignore[attr-defined]
    found = 0
    for idx, row in enumerate(rows):
        for cell in row.getElementsByType(TableCell):
            formula = cell.getAttribute("formula") or ""
            if "con.sql(sql)" not in formula:
                continue
            found += 1
            span = cell.getAttribute("numbercolumnsspanned")
            assert not span or int(span) <= 1, (idx, span, formula)
            empty = 0
            probe = idx + 1
            while probe < len(rows) and _ods_row_is_empty(rows[probe]):
                empty += 1
                probe += 1
            assert empty >= SQL_RESULTS_SPILL_GUTTER_ROWS, (idx, empty)
    assert found == 2


def test_build_ods_showcase_formulas_and_layout(tmp_path: Path) -> None:
    out_path = tmp_path / "python_showcase_demo.ods"
    build_ods_showcase(out_path)
    _assert_ods_formulas_and_layout(_ods_content_xml(out_path))


def test_shipped_ods_fixture_formulas_and_layout() -> None:
    _assert_ods_formulas_and_layout(_ods_content_xml(FIXTURE_ODS))
    from odf.opendocument import load
    from odf.table import Table

    doc = load(str(FIXTURE_ODS))
    sql_sheet = next(t for t in doc.spreadsheet.getElementsByType(Table) if t.getAttribute("name") == "SQL_DuckDB")
    _assert_ods_results_unmerged_with_clearance(sql_sheet)


def _assert_results_formula_is_short_and_quote_safe(formula: str, *, sql_range: str) -> str:
    """RESULTS =PY() is a short runner: SQL is a cell/range arg, not a giant string."""
    assert formula.startswith("="), formula
    payload = _quoted_formula_payload(formula)
    rest = formula[formula.rindex('"') + 1 :]
    assert len(payload) <= RESULTS_PY_CODE_MAX_LEN, (len(payload), payload)
    assert '"""' not in formula
    assert payload.count("'") <= 2  # only the table-name quotes in register('sales')
    assert "data[0]" in payload and "data[1]" in payload
    assert "con.sql(sql)" in payload
    assert "result=con.sql(sql).df()" in payload
    assert ".tolist()" not in payload
    assert sql_range in rest
    for marker in _SQL_EMBED_MARKERS:
        assert marker not in formula, marker
    # Premature string close made Calc treat SQL commas as OpenFormula
    # separators (Region, Category → region; category).
    assert "Region; Category" not in formula
    assert "SUM(Revenue);" not in formula
    return payload


def test_sql_results_gutter_covers_region_category_spill() -> None:
    """Sheet-only RESULTS need ≥15×5 empty (header + 12 Region×Category + clearance)."""
    assert SQL_RESULTS_SPILL_GUTTER_ROWS >= 15
    assert SQL_RESULTS_SPILL_GUTTER_COLS >= 5
    assert sql_results_gutter_rows("sheet_sales") == SQL_RESULTS_SPILL_GUTTER_ROWS
    assert sql_results_gutter_rows("sheet_marketing") == SQL_RESULTS_SPILL_GUTTER_ROWS
    assert sql_results_gutter_rows("join_zip") == 2


def _empty_rows_below_xlsx_results(ws: object) -> list[tuple[str, int]]:
    """(RESULTS coordinate, consecutive empty rows in column A until the next value)."""
    from openpyxl.utils import coordinate_to_tuple

    gaps: list[tuple[str, int]] = []
    for coord, unused_formula in _xlsx_sql_duckdb_result_formulas(ws):
        row = coordinate_to_tuple(coord)[0]  # (row, column), 1-based
        empty = 0
        probe = row + 1
        while probe <= ws.max_row:  # type: ignore[attr-defined]
            val = ws[f"A{probe}"].value  # type: ignore[index]
            if val not in (None, ""):
                break
            empty += 1
            probe += 1
        gaps.append((coord, empty))
    return gaps


def _xlsx_merge_overlaps(ws: object, min_row: int, max_row: int, min_col: int, max_col: int) -> object | None:
    """Return a merged range that intersects the rectangle, else None."""
    for rng in ws.merged_cells.ranges:  # type: ignore[attr-defined]
        if rng.max_row < min_row or rng.min_row > max_row:
            continue
        if rng.max_col < min_col or rng.min_col > max_col:
            continue
        return rng
    return None


def _assert_xlsx_results_unmerged_with_clearance(ws: object) -> None:
    """Live RESULTS origin is unmerged; 15×5 under it (plus B–E of the origin row) is empty."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import coordinate_to_tuple

    formulas = _xlsx_sql_duckdb_result_formulas(ws)
    assert len(formulas) == 2
    for coord, unused_formula in formulas:
        row, col = coordinate_to_tuple(coord)
        assert col == 1, coord
        origin_merge = _xlsx_merge_overlaps(ws, row, row, col, col)
        assert origin_merge is None, (coord, origin_merge)
        # Origin row B–E plus 15 rows × 5 cols under the formula.
        overlap = _xlsx_merge_overlaps(
            ws,
            row,
            row + SQL_RESULTS_SPILL_GUTTER_ROWS,
            1,
            SQL_RESULTS_SPILL_GUTTER_COLS,
        )
        assert overlap is None, (coord, overlap)
        for r in range(row, row + 1 + SQL_RESULTS_SPILL_GUTTER_ROWS):
            for c in range(1, SQL_RESULTS_SPILL_GUTTER_COLS + 1):
                if r == row and c == col:
                    continue
                cell = ws.cell(row=r, column=c)  # type: ignore[attr-defined]
                assert not isinstance(cell, MergedCell), cell.coordinate
                assert cell.value in (None, ""), (cell.coordinate, cell.value)


def test_generated_xlsx_results_have_spill_gutter(tmp_path: Path) -> None:
    """A 13-row spill from A19 must not hit the next section title."""
    out = tmp_path / "python_showcase_demo.xlsx"
    build_xlsx_showcase(out)
    from openpyxl import load_workbook

    ws = load_workbook(out)["SQL_DuckDB"]
    gaps = _empty_rows_below_xlsx_results(ws)
    assert len(gaps) == 2
    for coord, empty in gaps:
        assert empty >= SQL_RESULTS_SPILL_GUTTER_ROWS, (coord, empty)
    _assert_xlsx_results_unmerged_with_clearance(ws)


def test_sheet_only_result_formulas_are_short_and_read_sql_from_cell_arg() -> None:
    """SQL is not embedded; RESULTS passes an explicit SQL cell/range plus the data range."""
    # Ranges here are formula-builder examples (scenario 1 SQL stays A11:A16).
    cases = (
        ("sheet_sales", "A11:A16"),
        ("sheet_marketing", "A23:A29"),
    )
    for kind, sql_range in cases:
        for ods_fmt in (False, True):
            formula = _scenario_result_formula(kind, sql_range, ods=ods_fmt)
            assert formula is not None, kind
            _assert_results_formula_is_short_and_quote_safe(formula, sql_range=sql_range)
    assert _scenario_result_formula("join_zip", "A40:A54", ods=False) is None


def test_duckdb_sql_from_cell_code_is_quote_safe_and_under_cap() -> None:
    for table in ("sales", "marketing"):
        code = duckdb_sql_from_cell_code(table)
        assert len(code) <= RESULTS_PY_CODE_MAX_LEN
        assert '"' not in code
        assert f"register('{table}'" in code
        assert "data[1]" in code
        assert "result=con.sql(sql).df()" in code
        assert ".tolist()" not in code


def test_ods_formula_sql_results_two_args_not_rematched() -> None:
    formula = _scenario_result_formula("sheet_sales", "A11:A16", ods=True)
    assert formula is not None
    out = ods_formula(formula)
    assert "[$Sales_Analytics.A5:.J40]" in out
    assert "[.A11:.A16]" in out
    assert _NESTED_SHEET_REF.search(out) is None
    assert out.startswith(f"of:={CALC_PYTHON_ADDIN_FN}(")
    for marker in _SQL_EMBED_MARKERS:
        assert marker not in out


def test_sql_query_lines_keeps_visible_sql_out_of_the_formula() -> None:
    lines = sql_query_lines(SQL_SALES_BY_REGION_CATEGORY)
    assert any("SUM(Revenue)" in line for line in lines)
    assert len(lines) >= 4
    formula = _scenario_result_formula("sheet_sales", "A11:A16", ods=False)
    assert formula is not None
    assert "SUM(Revenue)" not in formula


def _xlsx_sql_duckdb_result_formulas(ws: object) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    for row in ws.iter_rows():  # type: ignore[attr-defined]
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("=") and "data[1]" in val:
                found.append((cell.coordinate, val))
    return found


def test_fixture_xlsx_sql_results_formulas_are_short_and_quote_safe() -> None:
    """Committed showcase xlsx RESULTS cells stay short; SQL lives in cells."""
    from openpyxl import load_workbook

    path = Path(__file__).resolve().parents[1] / "fixtures" / "python_showcase_demo.xlsx"
    wb = load_workbook(path)
    ws = wb["SQL_DuckDB"]
    formulas = _xlsx_sql_duckdb_result_formulas(ws)
    assert len(formulas) == 2
    sql_text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if isinstance(cell.value, str)
    )
    assert "SUM(Revenue)" in sql_text
    assert "SUM(Ad_Spend)" in sql_text
    assert "zip_income" in sql_text
    for unused_coord, formula in formulas:
        # Trailing arg is the SQL cell/range on this sheet (sales stays A11:A16;
        # marketing sits below the RESULTS spill gutter).
        rest = formula[formula.rindex('"') + 1 :]
        sql_arg = rest.rsplit(",", 1)[-1].strip().rstrip(")")
        _assert_results_formula_is_short_and_quote_safe(formula, sql_range=sql_arg)
    gaps = _empty_rows_below_xlsx_results(ws)
    assert len(gaps) == 2
    for coord, empty in gaps:
        assert empty >= SQL_RESULTS_SPILL_GUTTER_ROWS, (coord, empty)
    _assert_xlsx_results_unmerged_with_clearance(ws)
    join_notes = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and "query_folder_sql" in cell.value and not cell.value.startswith("=")
    ]
    assert join_notes, "ZIP join must stay a query_folder_sql pointer, not a live =PY()"
