# SPDX-License-Identifier: GPL-3.0-or-later
"""Orchestrate Excel ↔ DAG-style ``=PY`` conversion. Details in ``to_dag.py``."""

from __future__ import annotations

import json
import logging
import zipfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

from plugin.calc.excel_py_convert.parse_dag_formulas import iter_dag_py_formulas_xlsx
from plugin.calc.excel_py_convert.parse_excel_ooxml import load_excel_model
from plugin.calc.excel_py_convert.script_bank import iter_a1_span
from plugin.calc.excel_py_convert.to_dag import convert_model_to_dag
from plugin.calc.excel_py_convert.to_excel import convert_dag_cells_to_excel

if TYPE_CHECKING:
    from plugin.calc.excel_py_convert.models import ConversionReport, ConvertedCell

log = logging.getLogger(__name__)


def convert_to_dag(path: str | Path, *, best_effort: bool = False) -> ConversionReport:
    """Excel XLSX or JSON fixture → DAG-style conversion report."""
    model = load_excel_model(path)
    return convert_model_to_dag(model, best_effort=best_effort)


def convert_to_excel(path: str | Path) -> ConversionReport:
    """Workbook with DAG ``=PY`` formulas → Excel-shaped script/dependency export.

    This is **not** a full inverse that writes native ``pythonScripts.xml`` /
    ``_xlws.PY``; it reconstructs ``xl(%Pn%)`` script text and a display
    ``=PY("…", returnType)`` string from DAG formulas / report metadata.
    """
    path = Path(path)
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return convert_dag_cells_to_excel(_triples_from_json(data), report_meta=data if isinstance(data, dict) else None)
    triples = iter_dag_py_formulas_xlsx(path)
    report = convert_dag_cells_to_excel(triples)
    report.source_path = str(path)
    return report


def _triples_from_json(data: Any) -> list[tuple[str, str, str, dict[str, Any]]]:
    if isinstance(data, dict) and "cells" in data:
        out: list[tuple[str, str, str, dict[str, Any]]] = []
        for c in data["cells"]:
            formula = c.get("dag_formula") or c.get("formula")
            if not formula:
                continue
            out.append((str(c.get("sheet", "Sheet1")), str(c.get("cell", "A1")), str(formula), dict(c)))
        return out
    if isinstance(data, list):
        return [(str(x["sheet"]), str(x["cell"]), str(x["formula"]), dict(x)) for x in data]
    raise ValueError("JSON must be a dag report or list of {sheet, cell, formula}")


def convert_path(
    path: str | Path,
    *,
    direction: str,
    out_report: str | Path | None = None,
    best_effort: bool = False,
) -> ConversionReport:
    """Convert *path* in *direction* ``dag`` or ``excel``; optionally write JSON report."""
    direction = direction.strip().lower()
    if direction == "dag":
        report = convert_to_dag(path, best_effort=best_effort)
    elif direction == "excel":
        report = convert_to_excel(path)
    else:
        raise ValueError("direction must be 'dag' or 'excel'")
    if out_report is not None:
        Path(out_report).write_text(json.dumps(report.to_dict(), indent=2) + "\n", encoding="utf-8")
    return report


def _xlsx_formula_for_cell(cell: ConvertedCell) -> str:
    """Render a comma-separated OOXML ``=PY`` formula (script bank ref, no Calc sanitizer)."""
    from plugin.calc.excel_py_convert.script_bank import formula_for_converted_cell

    return formula_for_converted_cell(cell, separator=",", excel_escape=True, use_script_bank=True)


def _clear_spill_range(ws: Any, anchor: str, array_ref: str) -> None:
    """Clear cached/array result cells in *array_ref*, keeping the anchor for rewrite."""
    if not array_ref:
        return
    cells = iter_a1_span(array_ref)
    if len(cells) <= 1:
        return
    for coord in cells:
        if coord == anchor:
            continue
        try:
            ws[coord].value = None
        except Exception:
            continue


def _strip_python_in_excel_parts(out_path: Path) -> None:
    """Remove obsolete Python-in-Excel package parts after formula rewrite."""
    drop_prefixes = (
        "xl/pythonScripts",
        "xl/python",
    )
    drop_exact = {
        "xl/pythonScripts.xml",
    }
    tmp = out_path.with_suffix(out_path.suffix + ".tmpstrip")
    with zipfile.ZipFile(out_path, "r") as zin, zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            name = info.filename
            if name in drop_exact or any(name.startswith(p) for p in drop_prefixes):
                continue
            # Drop content-type / rel entries are left; orphan rels are harmless enough.
            # Controlled: also strip workbook rel targeting pythonScripts.
            data = zin.read(name)
            if name.endswith(".rels") or name == "[Content_Types].xml":
                text = data.decode("utf-8", errors="ignore")
                if "pythonScripts" in text or ("python" in text.lower() and "Override" in text):
                    # Remove lines referencing pythonScripts
                    lines = []
                    for line in text.splitlines(keepends=True):
                        if "pythonScripts" in line or "pythonScript" in line:
                            continue
                        if 'PartName="/xl/python' in line:
                            continue
                        lines.append(line)
                    data = "".join(lines).encode("utf-8")
            zout.writestr(info, data)
    tmp.replace(out_path)


def write_dag_formulas_xlsx(
    source_xlsx: str | Path,
    report: ConversionReport,
    out_path: str | Path,
    *,
    strip_python_parts: bool = True,
) -> None:
    """Copy *source_xlsx* and replace successfully converted PY cells with DAG formulas.

    - Parks rewritten Python on visible ``py_code_<Sheet>`` sheets at the **same A1**
      as each caller when ``len(code) > 1000``; shorter scripts stay inline in ``=PY("…")``.
    - OOXML formulas use **comma** separators (not Calc ``;``).
    - Clears the source array/spill ``ref`` range (except the anchor) so old
      cached results do not block the new spill.
    - Fails closed on unmapped sheet titles (no silent first-sheet fallback).
    - Optionally strips ``xl/pythonScripts.xml`` and related package parts.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import IllegalCharacterError

    from plugin.calc.excel_py_convert.script_bank import (
        collect_script_bank,
        report_safety_warnings,
        write_script_bank_openpyxl,
    )

    source_xlsx = Path(source_xlsx)
    out_path = Path(out_path)
    wb = load_workbook(source_xlsx)
    sheet_by_key = {ws.title: ws for ws in wb.worksheets}
    errors: list[str] = []

    bank, bank_warnings = collect_script_bank(report)
    for w in bank_warnings:
        log.warning("excel_py convert: %s", w)
    for w in report_safety_warnings(report):
        log.warning("excel_py convert: %s", w)
    write_script_bank_openpyxl(wb, bank)

    for cell in report.cells:
        if not cell.converted or not cell.converted_code:
            continue
        ws = sheet_by_key.get(cell.sheet)
        if ws is None:
            # Accept sheet1 → first sheet only when the report used fixture aliases
            # AND there is exactly one worksheet — still prefer exact titles.
            lower = {t.lower(): w for t, w in sheet_by_key.items()}
            ws = lower.get(cell.sheet.lower())
        if ws is None:
            errors.append(f"unmapped sheet {cell.sheet!r} for cell {cell.cell}")
            continue
        if cell.array_ref:
            _clear_spill_range(ws, cell.cell, cell.array_ref)
        formula = _xlsx_formula_for_cell(cell)
        try:
            ws[cell.cell] = formula
        except IllegalCharacterError as exc:
            errors.append(f"{cell.sheet}!{cell.cell}: {exc}")

    if errors:
        wb.close()
        raise ValueError("write_dag_formulas_xlsx failed:\n" + "\n".join(errors))

    wb.save(out_path)
    wb.close()
    if strip_python_parts:
        _strip_python_in_excel_parts(out_path)
