# SPDX-License-Identifier: GPL-3.0-or-later
"""Parse Excel OOXML for Python scripts, PY formulas, tables, and array ranges.

Uses namespace-aware ElementTree (not regex scans) so prefixed namespaces, shared
formulas, and XML entities are handled correctly.
"""

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from plugin.calc.excel_py_convert.models import ExcelPyCell, ExcelWorkbookModel, SheetInfo

_NS_OD_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

_RE_A1_CELL = re.compile(r"^\$?([A-Za-z]+)\$?(\d+)$")
_RE_XLWS_PY = re.compile(
    r"(?:_xlfn\.)?_xlws\.PY\s*\(\s*(\d+)\s*,\s*(\d+)(.*)\)$",
    re.IGNORECASE | re.DOTALL,
)


def _local(tag: str) -> str:
    if tag.startswith("{"):
        return tag.rsplit("}", 1)[-1]
    return tag


def _findall(parent: ET.Element, name: str) -> list[ET.Element]:
    return [el for el in parent.iter() if _local(el.tag) == name]


def _find_child(parent: ET.Element, name: str) -> ET.Element | None:
    for child in list(parent):
        if _local(child.tag) == name:
            return child
    return None


def _col_row(a1: str) -> tuple[int, int]:
    m = _RE_A1_CELL.match(a1.replace("$", "").strip())
    if not m:
        return 0, 0
    col_s, row_s = m.group(1).upper(), m.group(2)
    col = 0
    for ch in col_s:
        col = col * 26 + (ord(ch) - 64)
    return int(row_s), col


def _unescape_xml(text: str) -> str:
    return (
        text.replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&apos;", "'")
        .replace("&amp;", "&")
    )


def split_top_level_args(tail: str) -> list[str]:
    """Split comma-separated formula args with quote/paren awareness."""
    s = (tail or "").strip()
    if s.startswith(","):
        s = s[1:].strip()
    if not s:
        return []
    args: list[str] = []
    buf: list[str] = []
    depth = 0
    in_sq = False
    in_dq = False
    i = 0
    while i < len(s):
        ch = s[i]
        if in_sq:
            buf.append(ch)
            if ch == "'" and i + 1 < len(s) and s[i + 1] == "'":
                buf.append(s[i + 1])
                i += 2
                continue
            if ch == "'":
                in_sq = False
            i += 1
            continue
        if in_dq:
            buf.append(ch)
            if ch == '"':
                in_dq = False
            i += 1
            continue
        if ch == "'":
            in_sq = True
            buf.append(ch)
        elif ch == '"':
            in_dq = True
            buf.append(ch)
        elif ch == "(":
            depth += 1
            buf.append(ch)
        elif ch == ")":
            depth = max(0, depth - 1)
            buf.append(ch)
        elif ch == "," and depth == 0:
            args.append("".join(buf).strip())
            buf = []
        else:
            buf.append(ch)
        i += 1
    if buf:
        args.append("".join(buf).strip())
    return [a for a in args if a]


def parse_xlws_py_formula(formula: str) -> tuple[int, int, list[str]] | None:
    """Parse ``_xlfn._xlws.PY(scriptIndex, returnType, …deps)``."""
    text = (formula or "").strip()
    if text.startswith("="):
        text = text[1:].strip()
    m = _RE_XLWS_PY.match(text)
    if not m:
        return None
    script_index = int(m.group(1))
    return_type = int(m.group(2))
    deps = split_top_level_args(m.group(3) or "")
    return script_index, return_type, deps


def _resolve_rel_target(rels_path: str, target: str) -> str:
    """Resolve a Relationship Target to a package-relative path (forward slashes)."""
    target = target.replace("\\", "/").lstrip("/")
    if target.startswith("xl/") or target.startswith("[Content_Types"):
        return target
    # rels live next to the owning part: xl/worksheets/_rels/sheet1.xml.rels
    # owning dir is parent of _rels: xl/worksheets
    owning_dir = str(Path(rels_path).parent.parent).replace("\\", "/")
    if owning_dir == ".":
        owning_dir = ""
    joined = f"{owning_dir}/{target}" if owning_dir else target
    parts: list[str] = []
    for p in joined.replace("\\", "/").split("/"):
        if p == "..":
            if parts:
                parts.pop()
        elif p and p != ".":
            parts.append(p)
    return "/".join(parts)


def _parse_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    """Return relationship Id → package-relative Target."""
    out: dict[str, str] = {}
    try:
        root = ET.fromstring(zf.read(rels_path))
    except KeyError:
        return out
    for rel in root:
        if _local(rel.tag) != "Relationship":
            continue
        rid = rel.attrib.get("Id") or ""
        target = rel.attrib.get("Target") or ""
        if not rid or not target:
            continue
        out[rid] = _resolve_rel_target(rels_path, target)
    return out


def _rel_type_is_table(rel_type: str) -> bool:
    return rel_type.rstrip("/").endswith("table")


def _parse_rels_with_types(zf: zipfile.ZipFile, rels_path: str) -> list[tuple[str, str, str]]:
    """Return (Id, Type, Target) triples."""
    out: list[tuple[str, str, str]] = []
    try:
        root = ET.fromstring(zf.read(rels_path))
    except KeyError:
        return out
    for rel in root:
        if _local(rel.tag) != "Relationship":
            continue
        rid = rel.attrib.get("Id") or ""
        rtype = rel.attrib.get("Type") or ""
        target = rel.attrib.get("Target") or ""
        if not rid or not target:
            continue
        out.append((rid, rtype, _resolve_rel_target(rels_path, target)))
    return out


def _workbook_sheets(zf: zipfile.ZipFile) -> list[SheetInfo]:
    """Map workbook sheet order/titles to worksheet part paths."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = _parse_rels(zf, "xl/_rels/workbook.xml.rels")
    sheets_el = None
    for el in wb:
        if _local(el.tag) == "sheets":
            sheets_el = el
            break
    if sheets_el is None:
        return []
    out: list[SheetInfo] = []
    for order, sh in enumerate(list(sheets_el)):
        if _local(sh.tag) != "sheet":
            continue
        title = sh.attrib.get("name") or f"Sheet{order + 1}"
        rid = sh.attrib.get(f"{{{_NS_OD_REL}}}id") or ""
        if not rid:
            for k, v in sh.attrib.items():
                if k.endswith("}id") or k in ("r:id", "id"):
                    rid = v
                    break
        part = rels.get(rid, "")
        if part and not part.startswith("xl/"):
            part = f"xl/{part}"
        out.append(SheetInfo(title=title, order=order, part_name=part))
    return out


def _parse_python_scripts(zf: zipfile.ZipFile) -> list[str]:
    """Parse ``xl/pythonScripts.xml``.

    Microsoft stores scripts as ordered ``<pythonScript><code>…</code></pythonScript>``
    children (document order = script index). Some builds may also set an ``index`` attr.
    """
    try:
        raw = zf.read("xl/pythonScripts.xml")
    except KeyError:
        return []
    root = ET.fromstring(raw)
    # Direct pythonScript children in document order (ignore nested noise).
    script_els = [el for el in list(root) if _local(el.tag) == "pythonScript"]
    if not script_els:
        script_els = [el for el in root.iter() if _local(el.tag) == "pythonScript"]

    indexed: list[tuple[int, str]] = []
    sequential: list[str] = []
    for order, el in enumerate(script_els):
        code_el = _find_child(el, "code")
        text = "".join(code_el.itertext()) if code_el is not None else "".join(el.itertext())
        idx_s = el.attrib.get("index") or el.attrib.get("scriptIndex") or ""
        if idx_s != "":
            try:
                indexed.append((int(idx_s), text))
                continue
            except ValueError:
                pass
        sequential.append(text)

    if indexed and not sequential:
        indexed.sort(key=lambda t: t[0])
        max_i = max(i for i, _ in indexed)
        out = [""] * (max_i + 1)
        for i, body in indexed:
            out[i] = body
        return out
    # Document order is the script bank index (Excel samples).
    return sequential if sequential else [t[1] for t in sorted(indexed, key=lambda t: t[0])]


def _sheet_rels_path(part_name: str) -> str:
    p = Path(part_name)
    return str(p.parent / "_rels" / f"{p.name}.rels").replace("\\", "/")


def _qualify_sheet_ref(sheet_title: str, ref: str) -> str:
    """Return Sheet!A1:B2 with quoting when the sheet name needs it."""
    ref = ref.replace("$", "")
    if "!" in ref:
        return ref
    if re.search(r"[^\w]", sheet_title) or sheet_title[:1].isdigit():
        q = "'" + sheet_title.replace("'", "''") + "'"
    else:
        q = sheet_title
    return f"{q}!{ref}"


def _parse_table_ref(zf: zipfile.ZipFile, table_part: str, sheet_title: str) -> tuple[str, str] | None:
    try:
        root = ET.fromstring(zf.read(table_part))
    except KeyError:
        return None
    name = root.attrib.get("displayName") or root.attrib.get("name") or ""
    ref = root.attrib.get("ref") or ""
    if not name or not ref:
        return None
    return name, _qualify_sheet_ref(sheet_title, ref)


def _collect_array_refs(ws_root: ET.Element, sheet_title: str) -> dict[str, str]:
    """Map Sheet!Anchor → full array ref range from worksheet formula/@ref."""
    out: dict[str, str] = {}
    for c in _findall(ws_root, "c"):
        cell_ref = c.attrib.get("r") or ""
        f = _find_child(c, "f")
        if f is None or not cell_ref:
            continue
        arr = (f.attrib.get("ref") or "").strip().replace("$", "")
        if not arr:
            continue
        top_left = arr.split(":", 1)[0]
        out[_qualify_sheet_ref(sheet_title, top_left)] = arr if "!" in arr else _qualify_sheet_ref(sheet_title, arr)
        # Bare keys for fixtures / same-sheet lookups (first wins).
        out.setdefault(top_left, arr)
        out.setdefault(cell_ref.replace("$", ""), arr)
        out[_qualify_sheet_ref(sheet_title, cell_ref)] = out[_qualify_sheet_ref(sheet_title, top_left)]
    return out


def _shared_formula_map(ws_root: ET.Element) -> dict[str, str]:
    """Resolve shared formula masters (si → formula text)."""
    masters: dict[str, str] = {}
    for c in _findall(ws_root, "c"):
        f = _find_child(c, "f")
        if f is None:
            continue
        if (f.attrib.get("t") or "") != "shared":
            continue
        si = f.attrib.get("si")
        body = "".join(f.itertext()).strip()
        if si is not None and body:
            masters[si] = _unescape_xml(body)
    return masters


def _iter_py_cells(ws_root: ET.Element, sheet_title: str) -> list[ExcelPyCell]:
    masters = _shared_formula_map(ws_root)
    cells: list[ExcelPyCell] = []
    for c in _findall(ws_root, "c"):
        a1 = c.attrib.get("r") or ""
        f = _find_child(c, "f")
        if f is None or not a1:
            continue
        body = "".join(f.itertext()).strip()
        if (f.attrib.get("t") or "") == "shared" and not body:
            si = f.attrib.get("si")
            body = masters.get(si or "", "")
        if not body:
            continue
        formula = _unescape_xml(body)
        if "_xlws.PY" not in formula and "_xlws.py" not in formula.lower():
            continue
        parsed = parse_xlws_py_formula(formula)
        if parsed is None:
            continue
        script_index, return_type, deps = parsed
        row, col = _col_row(a1)
        cells.append(
            ExcelPyCell(
                sheet=sheet_title,
                cell=a1,
                script_index=script_index,
                return_type=return_type,
                deps=deps,
                formula_raw=formula if formula.startswith("=") else f"={formula}",
                array_ref=(f.attrib.get("ref") or "").replace("$", ""),
                row=row,
                col=col,
            )
        )
    return cells


def parse_excel_xlsx(path: str | Path) -> ExcelWorkbookModel:
    """Parse an ``.xlsx`` into scripts, PY cells, sheet map, tables, array anchors."""
    path = Path(path)
    with zipfile.ZipFile(path, "r") as zf:
        sheets = _workbook_sheets(zf)
        scripts = _parse_python_scripts(zf)
        tables: dict[str, str] = {}
        anchors: dict[str, str] = {}
        cells: list[ExcelPyCell] = []
        for sh in sheets:
            if not sh.part_name:
                continue
            try:
                ws_root = ET.fromstring(zf.read(sh.part_name))
            except KeyError:
                continue
            cells.extend(_iter_py_cells(ws_root, sh.title))
            anchors.update(_collect_array_refs(ws_root, sh.title))
            for _rid, rtype, target in _parse_rels_with_types(zf, _sheet_rels_path(sh.part_name)):
                if not _rel_type_is_table(rtype) and "tables/table" not in target:
                    continue
                parsed = _parse_table_ref(zf, target, sh.title)
                if parsed:
                    name, qref = parsed
                    tables[name] = qref
        for c in cells:
            if not c.row or not c.col:
                c.row, c.col = _col_row(c.cell)
        return ExcelWorkbookModel(
            scripts=scripts,
            cells=cells,
            sheets=sheets,
            tables=tables,
            anchor_snapshots=anchors,
            source_path=str(path),
        )


def _enrich_anchors_openpyxl(model: ExcelWorkbookModel, path: Path) -> None:
    try:
        import openpyxl
    except ImportError:
        return
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        for order, name in enumerate(wb.sheetnames):
            if not any(s.title == name for s in model.sheets):
                model.sheets.append(SheetInfo(title=name, order=order, part_name=""))
            ws = wb[name]
            af = getattr(ws, "array_formulae", None) or {}
            for anchor, meta in af.items():
                ref = getattr(meta, "ref", None) or str(meta)
                if not ref:
                    continue
                cleaned = str(ref).replace("$", "")
                model.anchor_snapshots[_qualify_sheet_ref(name, str(anchor))] = (
                    cleaned if "!" in cleaned else _qualify_sheet_ref(name, cleaned)
                )
                model.anchor_snapshots.setdefault(str(anchor), cleaned)
    finally:
        wb.close()


def load_excel_model(path: str | Path, *, prefer_openpyxl_anchors: bool = True) -> ExcelWorkbookModel:
    """Load from ``.xlsx`` or a JSON fixture matching ``ExcelWorkbookModel.to_dict()``."""
    path = Path(path)
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        model = ExcelWorkbookModel.from_dict(data)
        model.source_path = model.source_path or str(path)
        # Fixtures often use sheet1/sheet2 — synthesize sheet order if missing.
        if not model.sheets:
            titles: list[str] = []
            for c in model.cells:
                if c.sheet not in titles:
                    titles.append(c.sheet)
            model.sheets = [SheetInfo(title=t, order=i, part_name="") for i, t in enumerate(titles)]
        for c in model.cells:
            if not c.row or not c.col:
                c.row, c.col = _col_row(c.cell)
        return model
    model = parse_excel_xlsx(path)
    if prefer_openpyxl_anchors:
        _enrich_anchors_openpyxl(model, path)
    return model


# Back-compat alias used in older call sites / docs.
load_workbook_model = load_excel_model


def dump_model_json(model: ExcelWorkbookModel) -> dict[str, Any]:
    return model.to_dict()
