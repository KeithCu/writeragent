# SPDX-License-Identifier: GPL-3.0-or-later
"""Scan a workbook for DAG-style ``=PY`` / ``=PYTHON`` formulas."""

from __future__ import annotations

from pathlib import Path

from plugin.calc.python.formula_edit import parse_python_formula


def iter_dag_py_formulas_xlsx(path: str | Path) -> list[tuple[str, str, str]]:
    """Return ``(sheet_title, cell, formula)`` for PY/PYTHON cells in an ``.xlsx``."""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=False, read_only=True)
    out: list[tuple[str, str, str]] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    if parse_python_formula(val) is None:
                        continue
                    out.append((ws.title, cell.coordinate, val))
    finally:
        wb.close()
    return out
